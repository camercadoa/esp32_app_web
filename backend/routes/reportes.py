from flask import Blueprint, request, jsonify, send_file
import openpyxl
from config.database import get_connection
from datetime import datetime
import pytz, io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

reportes_bp = Blueprint('reportes', __name__, url_prefix='/api/reportes')


# ------------------------------------------------------------
# üü¶ Funci√≥n para obtener acciones filtradas
# ------------------------------------------------------------
def obtener_acciones_filtradas(fecha_inicio, fecha_fin, dispositivo_id=None):
    conexion = get_connection()
    with conexion.cursor() as cursor:
        query = """
            SELECT a.id, u.username, d.nombre AS dispositivo, a.accion, a.resultado, a.comentario, a.fecha_hora
            FROM acciones a
            JOIN usuarios u ON a.usuario_id = u.id
            JOIN dispositivos d ON a.dispositivo_id = d.id
            WHERE a.fecha_hora BETWEEN %s AND %s
        """
        params = [fecha_inicio, fecha_fin]
        if dispositivo_id:
            query += " AND a.dispositivo_id = %s"
            params.append(dispositivo_id)
        query += " ORDER BY a.fecha_hora DESC"

        cursor.execute(query, tuple(params))
        acciones = cursor.fetchall()
    conexion.close()

    # Convertir fecha a hora local Colombia
    tz_col = pytz.timezone("America/Bogota")
    for a in acciones:
        if isinstance(a["fecha_hora"], datetime):
            a["fecha_hora"] = a["fecha_hora"].replace(tzinfo=pytz.UTC).astimezone(tz_col).strftime("%Y-%m-%d %H:%M:%S")

    return acciones


# ------------------------------------------------------------
# üìä Reporte en formato Excel
# ------------------------------------------------------------
@reportes_bp.route('/excel', methods=['GET'])
def generar_excel():
    try:
        fecha_inicio = request.args.get("inicio")
        fecha_fin = request.args.get("fin")
        dispositivo_id = request.args.get("dispositivo_id", type=int)

        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Debe especificar las fechas 'inicio' y 'fin'"}), 400

        acciones = obtener_acciones_filtradas(fecha_inicio, fecha_fin, dispositivo_id)

        # Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Acciones"

        # Encabezados
        headers = ["ID", "Usuario", "Dispositivo", "Acci√≥n", "Resultado", "Comentario", "Fecha y hora"]
        ws.append(headers)

        # Datos
        for a in acciones:
            ws.append([a["id"], a["username"], a["dispositivo"], a["accion"], a["resultado"], a["comentario"], a["fecha_hora"]])

        # Estilo b√°sico
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = f"reporte_acciones_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, as_attachment=True, download_name=nombre_archivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print("‚ùå Error en /api/reportes/excel:", e)
        return jsonify({"error": str(e)}), 500


# ------------------------------------------------------------
# üìÑ Reporte en formato PDF
# ------------------------------------------------------------
@reportes_bp.route('/pdf', methods=['GET'])
def generar_pdf():
    try:
        fecha_inicio = request.args.get("inicio")
        fecha_fin = request.args.get("fin")
        dispositivo_id = request.args.get("dispositivo_id", type=int)

        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Debe especificar las fechas 'inicio' y 'fin'"}), 400

        acciones = obtener_acciones_filtradas(fecha_inicio, fecha_fin, dispositivo_id)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # T√≠tulo
        titulo = Paragraph(f"<b>Reporte de Acciones</b><br/>Desde {fecha_inicio} hasta {fecha_fin}", styles['Title'])
        elements.append(titulo)
        elements.append(Spacer(1, 12))

        # Tabla
        data = [["ID", "Usuario", "Dispositivo", "Acci√≥n", "Resultado", "Comentario", "Fecha y hora"]]
        for a in acciones:
            data.append([
                a["id"], a["username"], a["dispositivo"], a["accion"],
                a["resultado"], a["comentario"] or "-", a["fecha_hora"]
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4B8BBE")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#EAF2F8")),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        nombre_archivo = f"reporte_acciones_{fecha_inicio}_a_{fecha_fin}.pdf"
        return send_file(buffer, as_attachment=True, download_name=nombre_archivo, mimetype='application/pdf')

    except Exception as e:
        print("‚ùå Error en /api/reportes/pdf:", e)
        return jsonify({"error": str(e)}), 500
